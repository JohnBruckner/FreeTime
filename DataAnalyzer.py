#import os #May be useful for future iterations.
import openpyxl #For operating with data in Excel.
import re #Regex.
import io
import string
import matplotlib.pyplot as plt


freqDict = {"0":0}


###Finds the words in a string and splits into a list of words.
###Regex magic.
def wordSplit(strg):
    pattern =  re.compile(r"[\w]+")
    wList = pattern.findall(strg)
    wList = [word.upper() for word in wList]
    return wList

def containsAlpha(strg):
    for i in strg:
        if not i in string.ascii_letters:
            return False
    return True

#@TODO: Temporary solution. Need better function.
def filterDict(dct):
    dct = {k: v for k, v in dct.items() if containsAlpha(k)}
    return dct




###Where the bulk of the work happens.
###Reads file line by line and counts word frequency.
def readFile(fileName):
    with  io.open(fileName, 'r', encoding='utf8') as iFile:
        for line in iFile:
            wordFreq(wordSplit(line))
    return None



###@TODO: Rewrite the function in such a way that it returns something
###       and global variables aren't used anymore.
###Updates global dictionary with the new words found in the list and the number
###of aparitions of words already in there.
def wordFreq(sentence):
    global freqDict
    for word in sentence:
        if word in freqDict:
            freqDict[word]+=1
        else:
            freqDict.update({word : 1})
    return None


###Simple function to transform a dictionary into a list of tuples.
def dictToTuple(dictionary):
    l = list(zip(dictionary.keys(), dictionary.values()))
    return l

def tupleToDict(tuples):
    d = dict((x, y) for x, y in tuples)
    return d


###A sort function for tuples. Format assumed is (String, Int) for purpose of this exercise.
###Can be generalised by using key = 0 && key = 1 instead.
###Current syntax is key = alpha to sort by alphabetical order
###and key = num to sort by number of aparitions.
def sortBy(*args, **kwargs):
    if kwargs["key"] == "alpha":
        args[0].sort(key = lambda tup: tup[0])
    elif kwargs["key"] == "num":
        args[0].sort(key = lambda tup: tup[1], reverse=True)
    else:
        print("Wrong key! Please try again")
        return None
    return args[0]

# def sortD(*args):
#     l = args[0].sort(key = lambda tup: tup[1])
#     return l


###Creates an Excel workbook and fills it with data from dictionary.
###Takes a string as a parameter, string is the final name of the workbook.
def updateWb(wbName, dct):
    #global freqDict
    wb = openpyxl.Workbook()
    ws = wb.active

    for i in range(1, len(dct)+1):
        ws.cell(row=i, column=1).value = list(dct.keys())[i-1]
        ws.cell(row=i, column=2).value = list(dct.values())[i-1]

    wb.save(wbName)
    return None


#@TODO: Add a class for dictionary processing
#@TODO: Add proper graphing functionality
def main():
    global freqDict

    inFile = input("Please input the name of the text file (.txt) \n")
    wFile = input("\nPlease input the name of the workbook in which you want to save the results \n")

    readFile(inFile)
    toSort = freqDict
    toSort = filterDict(toSort)
    toSort = dictToTuple(toSort)
    #toSort = toSort[:100]
    toSort = sortBy(toSort, key="num")
    toSort = tupleToDict(toSort)

    # names = toSort.keys()
    # values = toSort.values()
    #
    # plt.bar(names, values)
    # plt.show()

    #print(toSort)

    updateWb(wFile, toSort)

    return None

main()
